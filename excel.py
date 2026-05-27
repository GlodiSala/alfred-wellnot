from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Facture Mai 2026"

# Couleurs
bleu_wellnot = "054561"
teal_wellnot = "14b0bd"
gris_clair = "F2F2F2"
blanc = "FFFFFF"

# Largeurs colonnes
ws.column_dimensions['A'].width = 55
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Bordure fine
bordure = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ── TITRE ────────────────────────────────────────────────
ws.merge_cells('A1:D1')
titre = ws['A1']
titre.value = "RÉCAPITULATIF DES PRESTATIONS — MAI 2026"
titre.font = Font(bold=True, size=14, color=blanc)
titre.fill = PatternFill("solid", fgColor=bleu_wellnot)
titre.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# Sous-titre
ws.merge_cells('A2:D2')
sous_titre = ws['A2']
sous_titre.value = "Wellnot — Projet Alfred · Congrès des Notaires belges septembre 2026"
sous_titre.font = Font(italic=True, size=10, color=blanc)
sous_titre.fill = PatternFill("solid", fgColor=teal_wellnot)
sous_titre.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 20

# Ligne vide
ws.row_dimensions[3].height = 10

# ── EN-TÊTES ─────────────────────────────────────────────
entetes = ['Description', 'Heures', 'Taux horaire', 'Montant']
for col, entete in enumerate(entetes, 1):
    cell = ws.cell(row=4, column=col)
    cell.value = entete
    cell.font = Font(bold=True, color=blanc)
    cell.fill = PatternFill("solid", fgColor=bleu_wellnot)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = bordure
ws.row_dimensions[4].height = 20

# ── LIGNES DE DONNÉES ─────────────────────────────────────
donnees = [
    (
        "V1 — Conception, architecture, avatar Alfred animé,\nintelligence artificielle, voix",
        "12h", "35 €/h", None
    ),
    (
        "V2 — Intégration sur la plateforme Alfred réelle,\nnavigation autonome, script bilingue FR/NL",
        "13h", "35 €/h", None
    ),
    (
        "Frais APIs Google (voix + intelligence artificielle)",
        "—", "—", 15
    ),
]

for i, (desc, heures, taux, montant) in enumerate(donnees, 5):
    fill = PatternFill("solid", fgColor=gris_clair) if i % 2 == 0 else PatternFill("solid", fgColor=blanc)
    
    # Description
    cell_desc = ws.cell(row=i, column=1)
    cell_desc.value = desc
    cell_desc.fill = fill
    cell_desc.border = bordure
    cell_desc.alignment = Alignment(wrap_text=True, vertical='center')
    cell_desc.font = Font(size=10)
    ws.row_dimensions[i].height = 35

    # Heures
    cell_h = ws.cell(row=i, column=2)
    cell_h.value = heures
    cell_h.fill = fill
    cell_h.border = bordure
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.font = Font(size=10)

    # Taux
    cell_t = ws.cell(row=i, column=3)
    cell_t.value = taux
    cell_t.fill = fill
    cell_t.border = bordure
    cell_t.alignment = Alignment(horizontal='center', vertical='center')
    cell_t.font = Font(size=10)

    # Montant avec formule
    cell_m = ws.cell(row=i, column=4)
    if i == 5:
        cell_m.value = "=12*35"
    elif i == 6:
        cell_m.value = "=13*35"
    else:
        cell_m.value = 15
    cell_m.number_format = '#,##0 "€"'
    cell_m.fill = fill
    cell_m.border = bordure
    cell_m.alignment = Alignment(horizontal='right', vertical='center')
    cell_m.font = Font(size=10)

# Ligne vide
ws.row_dimensions[8].height = 10

# ── TOTAL ─────────────────────────────────────────────────
ws.merge_cells('A9:B9')
total_label = ws['A9']
total_label.value = "TOTAL"
total_label.font = Font(bold=True, size=12, color=blanc)
total_label.fill = PatternFill("solid", fgColor=bleu_wellnot)
total_label.alignment = Alignment(horizontal='center', vertical='center')
total_label.border = bordure

cell_h_total = ws['C9']
cell_h_total.value = "25h"
cell_h_total.font = Font(bold=True, size=12, color=blanc)
cell_h_total.fill = PatternFill("solid", fgColor=bleu_wellnot)
cell_h_total.alignment = Alignment(horizontal='center', vertical='center')
cell_h_total.border = bordure

cell_total = ws['D9']
cell_total.value = "=SUM(D5:D7)"
cell_total.number_format = '#,##0 "€"'
cell_total.font = Font(bold=True, size=12, color=blanc)
cell_total.fill = PatternFill("solid", fgColor=bleu_wellnot)
cell_total.alignment = Alignment(horizontal='right', vertical='center')
cell_total.border = bordure
ws.row_dimensions[9].height = 30

# ── INFOS ─────────────────────────────────────────────────
ws.row_dimensions[11].height = 15
infos = [
    ("Prestations réalisées par :", "Glodi Sala"),
    ("Pour :", "Wellnot — Jean-François Ghigny"),
    ("Période :", "Mai 2026"),
    ("Coordonnées bancaires :", "[Ton IBAN]"),
]

for i, (label, valeur) in enumerate(infos, 11):
    cell_label = ws.cell(row=i, column=1)
    cell_label.value = label
    cell_label.font = Font(bold=True, size=10, color=bleu_wellnot)
    cell_label.alignment = Alignment(vertical='center')

    cell_val = ws.cell(row=i, column=2)
    cell_val.value = valeur
    cell_val.font = Font(size=10)
    ws.merge_cells(f'B{i}:D{i}')
    ws.row_dimensions[i].height = 18

# ── SAUVEGARDE ────────────────────────────────────────────
wb.save("Facture_Mai_2026_Wellnot.xlsx")
print("Fichier genere : Facture_Mai_2026_Wellnot.xlsx")